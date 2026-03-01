"""Tests for Excel tools in mcp_server.py."""

import json
import os

import pytest
from openpyxl import load_workbook

from onit_office.mcp_server import (
    create_excel,
    add_excel_rows,
    read_excel,
    modify_excel_cells,
)


class TestCreateExcel:
    """Tests for create_excel()."""

    def test_basic_creation(self, tmp_data_dir):
        result = json.loads(create_excel(
            path=str(tmp_data_dir / "basic.xlsx"),
            headers=["Name", "Age"],
            rows=[["Alice", 30]],
        ))
        assert result["status"] == "success"
        assert result["row_count"] == 1
        assert result["headers"] == ["Name", "Age"]
        assert os.path.exists(result["excel_file"])

    def test_default_path(self, tmp_data_dir):
        result = json.loads(create_excel(headers=["Col1"]))
        assert result["status"] == "success"
        assert os.path.exists(result["excel_file"])

    def test_custom_sheet_name(self, tmp_data_dir):
        path = str(tmp_data_dir / "custom_sheet.xlsx")
        result = json.loads(create_excel(
            path=path, sheet_name="Data",
            headers=["X"], rows=[["1"]],
        ))
        assert result["sheet_name"] == "Data"
        wb = load_workbook(path)
        assert "Data" in wb.sheetnames

    def test_headers_json_string(self, tmp_data_dir):
        result = json.loads(create_excel(
            path=str(tmp_data_dir / "json.xlsx"),
            headers='["A", "B", "C"]',
        ))
        assert result["headers"] == ["A", "B", "C"]

    def test_rows_json_string(self, tmp_data_dir):
        result = json.loads(create_excel(
            path=str(tmp_data_dir / "json_rows.xlsx"),
            headers=["Name"],
            rows='[["Alice"], ["Bob"]]',
        ))
        assert result["row_count"] == 2

    def test_no_headers(self, tmp_data_dir):
        result = json.loads(create_excel(
            path=str(tmp_data_dir / "no_headers.xlsx"),
            rows=[["1", "2"], ["3", "4"]],
        ))
        assert result["status"] == "success"
        assert result["headers"] == []

    def test_no_rows(self, tmp_data_dir):
        result = json.loads(create_excel(
            path=str(tmp_data_dir / "no_rows.xlsx"),
            headers=["A", "B"],
        ))
        assert result["row_count"] == 0

    def test_auto_width(self, tmp_data_dir):
        path = str(tmp_data_dir / "width.xlsx")
        create_excel(
            path=path,
            headers=["Short", "A Very Long Header Name"],
            rows=[["x", "y"]],
            auto_width=True,
        )
        wb = load_workbook(path)
        ws = wb.active
        # Column B should be wider than column A
        assert ws.column_dimensions["B"].width > ws.column_dimensions["A"].width

    def test_header_styling(self, tmp_data_dir):
        path = str(tmp_data_dir / "styled.xlsx")
        create_excel(path=path, headers=["Name", "Age"])
        wb = load_workbook(path)
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold is True


class TestAddExcelRows:
    """Tests for add_excel_rows()."""

    def test_add_rows(self, xlsx_file):
        result = json.loads(add_excel_rows(
            path=xlsx_file,
            rows=[["Charlie", 35, "charlie@example.com"]],
        ))
        assert result["status"] == "success"
        assert result["rows_added"] == 1
        assert result["total_rows"] == 4  # header + 2 original + 1 new

    def test_add_multiple_rows(self, xlsx_file):
        result = json.loads(add_excel_rows(
            path=xlsx_file,
            rows=[["C", 1, "c@x.com"], ["D", 2, "d@x.com"], ["E", 3, "e@x.com"]],
        ))
        assert result["rows_added"] == 3

    def test_json_string_rows(self, xlsx_file):
        result = json.loads(add_excel_rows(
            path=xlsx_file,
            rows='[["Frank", 40, "frank@x.com"]]',
        ))
        assert result["rows_added"] == 1

    def test_specific_sheet(self, xlsx_file):
        result = json.loads(add_excel_rows(
            path=xlsx_file,
            rows=[["New", 1, "new@x.com"]],
            sheet_name="Sheet1",
        ))
        assert result["sheet_name"] == "Sheet1"

    def test_invalid_sheet_name(self, xlsx_file):
        result = json.loads(add_excel_rows(
            path=xlsx_file,
            rows=[["X", 1, "x@x.com"]],
            sheet_name="NonExistent",
        ))
        assert "error" in result

    def test_nonexistent_file(self, tmp_data_dir):
        result = json.loads(add_excel_rows(
            path=str(tmp_data_dir / "nope.xlsx"),
            rows=[["A", 1]],
        ))
        assert "error" in result


class TestReadExcel:
    """Tests for read_excel()."""

    def test_read_basic(self, xlsx_file):
        result = json.loads(read_excel(path=xlsx_file))
        assert result["status"] == "success"
        assert result["headers"] == ["Name", "Age", "Email"]
        assert result["row_count"] == 2
        assert result["total_rows"] == 2
        assert result["truncated"] is False

    def test_read_rows_content(self, xlsx_file):
        result = json.loads(read_excel(path=xlsx_file))
        rows = result["rows"]
        assert rows[0][0] == "Alice"
        assert rows[1][0] == "Bob"

    def test_max_rows(self, xlsx_file):
        # Add many rows
        add_excel_rows(
            path=xlsx_file,
            rows=[[f"User{i}", i, f"user{i}@x.com"] for i in range(10)],
        )
        result = json.loads(read_excel(path=xlsx_file, max_rows=5))
        assert result["row_count"] == 5
        assert result["total_rows"] == 12
        assert result["truncated"] is True

    def test_specific_sheet(self, xlsx_file):
        result = json.loads(read_excel(path=xlsx_file, sheet_name="Sheet1"))
        assert result["sheet_name"] == "Sheet1"

    def test_invalid_sheet(self, xlsx_file):
        result = json.loads(read_excel(path=xlsx_file, sheet_name="NoSheet"))
        assert "error" in result

    def test_nonexistent_file(self, tmp_data_dir):
        result = json.loads(read_excel(path=str(tmp_data_dir / "nope.xlsx")))
        assert "error" in result


class TestModifyExcelCells:
    """Tests for modify_excel_cells()."""

    def test_modify_single_cell(self, xlsx_file):
        result = json.loads(modify_excel_cells(
            path=xlsx_file,
            updates=[{"cell": "A2", "value": "Updated Alice"}],
        ))
        assert result["status"] == "success"
        assert result["cells_updated"] == 1

        wb = load_workbook(xlsx_file)
        assert wb.active["A2"].value == "Updated Alice"

    def test_modify_multiple_cells(self, xlsx_file):
        result = json.loads(modify_excel_cells(
            path=xlsx_file,
            updates=[
                {"cell": "A2", "value": "New Name"},
                {"cell": "B2", "value": 99},
                {"cell": "C2", "value": "new@x.com"},
            ],
        ))
        assert result["cells_updated"] == 3

    def test_json_string_updates(self, xlsx_file):
        result = json.loads(modify_excel_cells(
            path=xlsx_file,
            updates='[{"cell": "A2", "value": "JSON Update"}]',
        ))
        assert result["cells_updated"] == 1

    def test_no_updates(self, xlsx_file):
        result = json.loads(modify_excel_cells(path=xlsx_file))
        assert "error" in result

    def test_specific_sheet(self, xlsx_file):
        result = json.loads(modify_excel_cells(
            path=xlsx_file,
            updates=[{"cell": "A1", "value": "Header"}],
            sheet_name="Sheet1",
        ))
        assert result["sheet_name"] == "Sheet1"

    def test_invalid_sheet(self, xlsx_file):
        result = json.loads(modify_excel_cells(
            path=xlsx_file,
            updates=[{"cell": "A1", "value": "X"}],
            sheet_name="NoSheet",
        ))
        assert "error" in result

    def test_nonexistent_file(self, tmp_data_dir):
        result = json.loads(modify_excel_cells(
            path=str(tmp_data_dir / "nope.xlsx"),
            updates=[{"cell": "A1", "value": "X"}],
        ))
        assert "error" in result

    def test_update_with_missing_cell_key(self, xlsx_file):
        result = json.loads(modify_excel_cells(
            path=xlsx_file,
            updates=[{"value": "No Cell Key"}],
        ))
        # Should skip entries without "cell" key
        assert result["cells_updated"] == 0
